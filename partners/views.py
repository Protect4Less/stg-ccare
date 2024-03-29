from django.shortcuts import render, redirect
# from django.http import  HttpResponse
# from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from backend.dao.PartnersDAO import PartnersDAO
# from backend.dao.CampaignDAO import CampaignDAO
# from backend.dao.MasterDAO import MasterDAO
# from backend.lib.CommonLib import CommonLib
from django.contrib import messages
# from django.contrib.auth.hashers import make_password
# from django.http import JsonResponse
# from django.utils.encoding import smart_str
# import datetime
# import json
# from backend.dao.ClaimDAO import ClaimDAO
# import os
# from django.conf import settings
# from django.http import HttpResponse, Http404
# import xlrd
import openpyxl
# from django.db import connection
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import logging

logger = logging.getLogger(__name__)


@login_required(login_url='/login')
def upload_create_policy(request):
    response = {}
    error = ''
    excel_file = request.FILES.get('item_data_excel', None)
    partner_code = request.POST.get('partner_code', None)

    partners_config = [
        {"id": 1014, "name": '1014 - Florance'},
        {"id": 1025, "name": '1025 - SAFARI HYPER MARKET - SADIQ ALI'},
        {"id": 1026, "name": '1026 - NESTO GROUP - MR. FARHAN MOHAMED'},
        {"id": 'RG', "name": 'RG - Redington'},
        {"id": 1030, "name": '1030 - TECH-OFFER (FLORENCE TRD)'},
        {"id": 1031, "name": '1031 - THOMSUN PLAY'},
        {"id": 1040, "name": '1040 - AL NOOR AL ABYADH'},
        {"id": 1051, "name": '1051 - JACKYS RETAIL'},
        {"id": 1052, "name": '1052 - GAME OVER PLUS ELECTRONICS'},
        {"id": 1053, "name": '1053 - ASIA MOBILE PHONE LLC'},
        {"id": 1054, "name": '1054 - ABUDHABI COOP'},
        {"id": 1055, "name": '1055 - Al MALAKAH ALA ZAHABIYA LLC'},
        {"id": 1056, "name": '1056 - AL TAMAM ELECTRONICS'},
        {"id": 1057, "name": '1057 - Asia Palace Mobile Phones LLC'},
        {"id": 1059, "name": '1059 - LivLyt FZ LLC'},
        {"id": 1060, "name": '1060 - MASTERPIECE LLC'},
        {"id": 1061, "name": '1061 - Desert Beat Trading LLC'},
        {"id": 1062, "name": '1062 - GEANT HYPERMARKET'},
        {"id": 1064, "name": '1064 - SOL SUPER CLICK TRADING L.L.C'},
        {"id": 1065, "name": '1065 - ASIA MOBILE PHONE'},
        {"id": 1066, "name": '1066 - Denaster General Trading LLC'},
        {"id": 1067, "name": '1067 - P4LOFFLINE'},
        {"id": 1068, "name": '1068 - Fixsquad - Offline'},
        {"id": 1069, "name": '1069 - REVENT'},
        {"id": 1070, "name": '1070 - Fone Garage'},
        {"id": 1071, "name": '1071 - Papita Trading LLC'},
        {"id": 1072, "name": '1072 - COSTLESS HUB ELECTRONICS LLC'},
    ]
    try:
        if excel_file is not None and partner_code is not None:

            excel_file = request.FILES['item_data_excel']
            wb = openpyxl.load_workbook(excel_file)
            worksheet = wb["Sheet1"]
            print(worksheet)

            logger.debug(worksheet)

            excel_data = list()
            cnt = 0
            cntin = 0

            excel_data = list()
            # iterating over the rows and
            # getting value from each cell in row
            for row in worksheet.iter_rows():
                row_data = list()
                print("===================================================")
                for cell in row:
                    cell_coordinate = coordinate_from_string(cell.coordinate)
                    row_number = cell_coordinate[1]

                    if cell.value in ["", None, " "]:
                        continue

                    cell.value = cell.value.strip() if isinstance(cell.value, str) else cell.value

                    if cell.value in ["invoice_no", "Invoice Number"]:
                        invoice_no_coordinate = coordinate_from_string(cell.coordinate)
                        invoice_no_col = invoice_no_coordinate[0]

                    if cell.value in ["sku", "SKU ( Article )", "Part#", "Part"]:
                        sku_coordinate = coordinate_from_string(cell.coordinate)
                        sku_col = sku_coordinate[0]

                    if cell.value in ["location", "Location"]:
                        location_coordinate = coordinate_from_string(cell.coordinate)
                        location_col = location_coordinate[0]

                    if cell.value in ["device", "Device", "Category"]:
                        device_coordinate = coordinate_from_string(cell.coordinate)
                        device_col = device_coordinate[0]

                    if cell.value in ["sub_device", "Sub Device"]:
                        sub_device_coordinate = coordinate_from_string(cell.coordinate)
                        sub_device_col = sub_device_coordinate[0]

                    if cell.value in ["model", "Model", "Description", "Model Name"]:
                        model_coordinate = coordinate_from_string(cell.coordinate)
                        model_col = model_coordinate[0]

                    if cell.value in ["Model ID"]:
                        model_id_coordinate = coordinate_from_string(cell.coordinate)
                        model_id_col = model_id_coordinate[0]

                    if cell.value in ["brand", "Brand"]:
                        brand_coordinate = coordinate_from_string(cell.coordinate)
                        brand_col = brand_coordinate[0]

                    if cell.value in ["purchase_month", "Purchase Date", "Device Purchase Date"]:
                        purchase_month_coordinate = coordinate_from_string(cell.coordinate)
                        purchase_month_col = purchase_month_coordinate[0]

                    if cell.value in ["first_name", "First Name", "Customer Name"]:
                        first_name_coordinate = coordinate_from_string(cell.coordinate)
                        first_name_col = first_name_coordinate[0]

                    if cell.value in ["last_name", "Last Name"]:
                        last_name_coordinate = coordinate_from_string(cell.coordinate)
                        last_name_col = last_name_coordinate[0]

                    if cell.value in ["email", "Email ID"]:
                        email_coordinate = coordinate_from_string(cell.coordinate)
                        email_col = email_coordinate[0]

                    if cell.value in ["mobile_number", "Mobile Number"]:
                        mobile_number_coordinate = coordinate_from_string(cell.coordinate)
                        mobile_number_col = mobile_number_coordinate[0]

                    if cell.value in ["imei_serial_no", "Imei / Serial Nuber", "IMEI Number", "IMEI No.", "IMEI No"]:
                        imei_serial_no_coordinate = coordinate_from_string(cell.coordinate)
                        imei_serial_no_col = imei_serial_no_coordinate[0]

                    if cell.value in ["term_type", "Term Type"]:
                        term_type_coordinate = coordinate_from_string(cell.coordinate)
                        term_type_col = term_type_coordinate[0]

                    if cell.value in ["device_value", "Device Value"]:
                        invoice_value_coordinate = coordinate_from_string(cell.coordinate)
                        invoice_value_col = invoice_value_coordinate[0]

                    if cell.value in ["device_currency", "Device Currency"]:
                        device_currency_coordinate = coordinate_from_string(cell.coordinate)
                        device_currency_col = device_currency_coordinate[0]
                        print('device_currency_coordinate:: ', device_currency_coordinate)

                    if cell.value in ["Plan Activation Date", "Date"]:
                        plan_ativation_date_coordinate = coordinate_from_string(cell.coordinate)
                        plan_ativation_date_col = plan_ativation_date_coordinate[0]

                    if cell.value in ["Device Name"]:
                        device_name_coordinate = coordinate_from_string(cell.coordinate)
                        device_name_col = device_name_coordinate[0]

                    if cell.value in ["Plan Price"]:
                        plan_price_coordinate = coordinate_from_string(cell.coordinate)
                        plan_price_col = plan_price_coordinate[0]

                    if cell.value in ["Plan Tax"]:
                        plan_tax_coordinate = coordinate_from_string(cell.coordinate)
                        plan_tax_col = plan_tax_coordinate[0]

                    if cell.value in ["Plan Total Price"]:
                        plan_total_price_coordinate = coordinate_from_string(cell.coordinate)
                        plan_total_price_col = plan_total_price_coordinate[0]

                    if cell.value in ["XCQC Order id"]:
                        xcqc_id_coordinate = coordinate_from_string(cell.coordinate)
                        xcqc_id_col = xcqc_id_coordinate[0]

                if 'device_currency_col' in locals():
                    device_currency_cell = "{}{}".format(device_currency_col, row_number)
                    device_currency_value = str(worksheet[device_currency_cell].value)

                if 'invoice_value_col' in locals():
                    invoice_value_cell = "{}{}".format(invoice_value_col, row_number)
                    invoice_value_value = str(worksheet[invoice_value_cell].value)

                if 'location_col' in locals():
                    location_value_cell = "{}{}".format(location_col, row_number)
                    location_value = str(worksheet[location_value_cell].value)

                if 'term_type_col' in locals():
                    term_type_cell = "{}{}".format(term_type_col, row_number)
                    term_type_value = str(worksheet[term_type_cell].value)

                if 'imei_serial_no_col' in locals():
                    imei_serial_no_cell = "{}{}".format(imei_serial_no_col, row_number)
                    imei_serial_no_value = str(worksheet[imei_serial_no_cell].value)

                if 'mobile_number_col' in locals():
                    mobile_number_cell = "{}{}".format(mobile_number_col, row_number)
                    mobile_number_value = str(worksheet[mobile_number_cell].value)

                if 'email_col' in locals():
                    email_cell = "{}{}".format(email_col, row_number)
                    email_value = str(worksheet[email_cell].value)

                if 'last_name_col' in locals():
                    last_name_cell = "{}{}".format(last_name_col, row_number)
                    last_name_value = str(worksheet[last_name_cell].value)

                if 'first_name_col' in locals():
                    first_name_cell = "{}{}".format(first_name_col, row_number)
                    first_name_value = str(worksheet[first_name_cell].value)

                if 'purchase_month_col' in locals():
                    purchase_month_cell = "{}{}".format(purchase_month_col, row_number)
                    purchase_month_value = str(worksheet[purchase_month_cell].value)

                if 'model_col' in locals():
                    model_cell = "{}{}".format(model_col, row_number)
                    model_value = str(worksheet[model_cell].value)

                if 'model_id_col' in locals():
                    model_id_cell = "{}{}".format(model_id_col, row_number)
                    model_code_value = str(worksheet[model_id_cell].value)

                if 'device_col' in locals():
                    device_cell = "{}{}".format(device_col, row_number)
                    device_value = str(worksheet[device_cell].value)

                if 'sub_device_col' in locals():
                    sub_device_cell = "{}{}".format(sub_device_col, row_number)
                    sub_device_value = str(worksheet[sub_device_cell].value)

                if 'invoice_no_col' in locals():
                    invoice_no_cell = "{}{}".format(invoice_no_col, row_number)
                    invoice_no_value = str(worksheet[invoice_no_cell].value)

                if 'sku_col' in locals():
                    sku_cell = "{}{}".format(sku_col, row_number)
                    sku_value = str(worksheet[sku_cell].value)

                if 'brand_col' in locals():
                    brand_cell = "{}{}".format(brand_col, row_number)
                    brand_value = str(worksheet[brand_cell].value)

                if 'plan_ativation_date_col' in locals():
                    plan_ativation_date_cell = "{}{}".format(plan_ativation_date_col, row_number)
                    plan_ativation_date_value = str(worksheet[plan_ativation_date_cell].value)

                if 'device_name_col' in locals():
                    device_name_cell = "{}{}".format(device_name_col, row_number)
                    device_name_value = str(worksheet[device_name_cell].value)

                if 'plan_price_col' in locals():
                    plan_price_cell = "{}{}".format(plan_price_col, row_number)
                    plan_price_value = str(worksheet[plan_price_cell].value)

                if 'plan_tax_col' in locals():
                    plan_tax_cell = "{}{}".format(plan_tax_col, row_number)
                    plan_tax_value = str(worksheet[plan_tax_cell].value)

                if 'plan_total_price_col' in locals():
                    plan_total_price_cell = "{}{}".format(plan_total_price_col, row_number)
                    plan_total_price_value = str(worksheet[plan_total_price_cell].value)

                if 'xcqc_id_col' in locals():
                    xcqc_id_cell = "{}{}".format(xcqc_id_col, row_number)
                    xcqc_id_value = str(worksheet[xcqc_id_cell].value)

                if row_number != 1 and partner_code in [
                    '1026', '1030', '1031', '1025', '1014', '1038', '1033', '1041',
                    '1036', '1044', '1035', '1026', '1046', '1051', '1049', '1052',
                    '1040', '1053', '1054', '1055', '1056', '1057', '1059', '1060',
                    '1061', '1062', '1064', '1065', '1066', '1067', '1068', '1069',
                    '1070', '1072'
                ] and email_value != 'None':
                    sku_value = sku_value if sku_value is not None and sku_value != "None" else ""
                    popd_data = PartnersDAO.insert_partners_offline_policy_data({
                        'popd_partner_code': partner_code,
                        # 'popd_xcqc_id':xcqc_id_value if 'xcqc_id_value' in locals() else '',
                        'popd_invoice_no': invoice_no_value if 'invoice_no_value' in locals() else '',
                        'popd_invoice_value': invoice_value_value if 'invoice_value_value' in locals() else '',
                        'popd_plan_price': plan_price_value if 'plan_price_value' in locals() else '',
                        'popd_plan_tax': plan_tax_value if 'plan_tax_value' in locals() else '',
                        'popd_plan_total_price': plan_total_price_value if 'plan_total_price_value' in locals() else '',
                        'popd_sku': sku_value if 'sku_value' in locals() else '',
                        'popd_location': location_value if 'location_value' in locals() else '',
                        'popd_device': device_value if 'device_value' in locals() else '',
                        'popd_device_name': device_name_value if 'device_name_value' in locals() else '',
                        'popd_sub_device': sub_device_value if 'sub_device_value' in locals() else '',
                        'popd_brand': brand_value if 'brand_value' in locals() else '',
                        'popd_model': model_value if 'model_value' in locals() else '',
                        'popd_model_code': model_code_value if 'model_code_value' in locals() else '',
                        'popd_purchase_month': purchase_month_value if 'purchase_month_value' in locals() else '',
                        'popd_first_name': first_name_value if 'first_name_value' in locals() else '',
                        'popd_last_name': last_name_value if 'last_name_value' in locals() else '',
                        'popd_email': email_value if 'email_value' in locals() else '',
                        'popd_mobile_number': mobile_number_value if 'mobile_number_value' in locals() else '',
                        'popd_imei_serial_no': imei_serial_no_value if 'imei_serial_no_value' in locals() else '',
                        'popd_term_type': term_type_value if 'term_type_value' in locals() else '',
                        'popd_device_currency': device_currency_value if 'device_currency_value' in locals() else '',
                        'popd_activation_date': plan_ativation_date_value if 'plan_ativation_date_value' in locals() else '',
                    })

                    logger.debug(popd_data)

                if row_number != 1 and partner_code in ['RG'] and imei_serial_no_value != 'None':
                    PartnersDAO.insert_partners_redington_policy_data({
                        'prpd_partner_code': partner_code,
                        'prpd_device_name': model_value if 'model_value' in locals() else '',
                        'prpd_imei_serial_no': imei_serial_no_value if 'imei_serial_no_value' in locals() else '',
                        'prpd_device': 'Mobile Phone',
                        'prpd_brand': 'APPLE',
                        'prpd_model': model_value if 'model_value' in locals() else '',
                        'prpd_capacity': '',
                        'prpd_part_sku': sku_value if 'sku_value' in locals() else '',
                        'prpd_retailer_name': first_name_value if 'first_name_value' in locals() else '',
                        'prpd_invoice_dt': plan_ativation_date_value if 'plan_ativation_date_value' in locals() else '',
                    })

                if row_number != 1 and partner_code in ['1071'] and imei_serial_no_value != 'None':
                    PartnersDAO.insert_partners_papita_policy_data({
                        'pppd_partner_code': partner_code,
                        'pppd_invoice_no': invoice_no_value if 'invoice_no_value' in locals() else '',
                        'pppd_category': device_value if 'device_value' in locals() else '',
                        'pppd_brand': brand_value if 'brand_value' in locals() else '',
                        'pppd_model': model_value if 'model_value' in locals() else '',
                        'pppd_purchase_date': purchase_month_value if 'purchase_month_value' in locals() else '',
                        'pppd_imei_no': imei_serial_no_value if 'imei_serial_no_value' in locals() else '',
                        'pppd_term_type': term_type_value if 'term_type_value' in locals() else '',
                        'pppd_invoice_value': invoice_value_value if 'invoice_value_value' in locals() else '',
                        'pppd_status': 'pending',
                    })

                # DYS
                if row_number != 1 and partner_code in ['1043'] and imei_serial_no_value != 'None':
                    PartnersDAO.insert_partners_dys_policy_data({
                        'pdpd_partner_code': partner_code,
                        'pdpd_invoice_dt': plan_ativation_date_value if 'plan_ativation_date_value' in locals() else '',
                        'pdpd_invoice_no': invoice_no_value if 'invoice_no_value' in locals() else '',
                        'pdpd_invoice_value': invoice_value_value if 'invoice_value_value' in locals() else '',
                        'pdpd_plan_price': plan_price_value if 'plan_price_value' in locals() else '',
                        'pdpd_plan_tax': plan_tax_value if 'plan_tax_value' in locals() else '',
                        'pdpd_plan_total_price': plan_total_price_value if 'plan_total_price_value' in locals() else '',
                        'pdpd_part_sku': sku_value if 'sku_value' in locals() else '',
                        'pdpd_location': location_value if 'location_value' in locals() else '',
                        'pdpd_device': device_value if 'device_value' in locals() else '',
                        'pdpd_device_name': device_name_value if 'device_name_value' in locals() else '',
                        'pdpd_sub_device': sub_device_value if 'sub_device_value' in locals() else '',
                        'pdpd_brand': brand_value if 'brand_value' in locals() else '',
                        'pdpd_model': model_value if 'model_value' in locals() else '',
                        'pdpd_model_code': model_code_value if 'model_code_value' in locals() else '',
                        'pdpd_purchase_month': purchase_month_value if 'purchase_month_value' in locals() else '',
                        'pdpd_first_name': first_name_value if 'first_name_value' in locals() else '',
                        'pdpd_last_name': last_name_value if 'last_name_value' in locals() else '',
                        'pdpd_email': email_value if 'email_value' in locals() else '',
                        'pdpd_mobile_number': mobile_number_value if 'mobile_number_value' in locals() else '',
                        'pdpd_imei_serial_no': imei_serial_no_value if 'imei_serial_no_value' in locals() else '',
                        'pdpd_term_type': term_type_value if 'term_type_value' in locals() else '',
                        'pdpd_device_currency': device_currency_value if 'device_currency_value' in locals() else '',
                        'pdpd_activation_date': plan_ativation_date_value if 'plan_ativation_date_value' in locals() else '',
                    })

            messages.success(request, 'File Uploaded successfully. Data will be processed')
    except Exception as e:
        error = str(e)
        messages.error(request, error)

    template_name = 'partners/upload_create_policy.html'
    partners_obj = PartnersDAO.get_partners(condition={'partners_status': 'active'})
    context = {"partners_obj": partners_obj, 'partners_config': partners_config}
    return render(request, template_name, context)


@login_required(login_url='/login')
def upload_renewal_policy(request):
    response = {}
    error = ''
    excel_file = request.FILES.get('item_data_excel', None)
    partner_code = request.POST.get('partner_code', None)

    policy_no_value = None
    term_type_value = None
    term_type_duration_value = None

    partners_config = [
        {"id": 1059, "name": '1059 - LivLyt FZ LLC'},
    ]

    if excel_file is not None and partner_code is not None:

        excel_file = request.FILES['item_data_excel']
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Sheet1"]

        excel_data = list()
        cnt = 0
        cntin = 0

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                cell_coordinate = coordinate_from_string(cell.coordinate)
                row_number = cell_coordinate[1]

                if cell.value in ["", None, " "]:
                    continue

                cell.value = cell.value.strip() if isinstance(cell.value, str) else cell.value

                if cell.value in ["old_policy_number", "old policy number"]:
                    policy_no_coordinate = coordinate_from_string(cell.coordinate)
                    policy_no_col = policy_no_coordinate[0]

                if cell.value in ["new_term_Type", "new term tpe"]:
                    term_type_coordinate = coordinate_from_string(cell.coordinate)
                    term_type_col = term_type_coordinate[0]

                if cell.value in ["new_term_type_duration_months"]:
                    term_type_duration_coordinate = coordinate_from_string(cell.coordinate)
                    term_type_duration_col = term_type_duration_coordinate[0]

            if 'policy_no_col' in locals():
                policy_no_cell = "{}{}".format(policy_no_col, row_number)
                policy_no_value = str(worksheet[policy_no_cell].value)

            if 'term_type_col' in locals():
                term_type_cell = "{}{}".format(term_type_col, row_number)
                term_type_value = str(worksheet[term_type_cell].value)

            if 'term_type_duration_col' in locals():
                term_type_duration_cell = "{}{}".format(term_type_duration_col, row_number)
                term_type_duration_value = str(worksheet[term_type_duration_cell].value)

            policy_data = PartnersDAO.get_user_policy(
                column='up_userid_id, up_s_id, up_partner_code',
                condition={'up_policy_no': policy_no_value}
            )
            if row_number != 1 and len(policy_data) == 0:
                error += f"Policy no. '{policy_no_value}' not found.<br>"

            # print()
            # print("policy_data \t:", policy_data)
            # print()

            if len(policy_data) > 0 and policy_data[0]['up_partner_code'] != "1059":
                error += f"Policy no. {policy_no_value} is not mapped with partner code 1059.<br>"

            if error == '':
                if row_number != 1 and partner_code in ['1059']:

                    renewal_data = PartnersDAO.insert_livlyt({
                        'lpr_u_id': policy_data[0]['up_userid_id'],
                        'lpr_s_id': policy_data[0]['up_s_id'],
                        'lpr_policy_no': policy_no_value if 'policy_no_col' in locals() else '',
                        'lpr_term_type_duration_months': term_type_duration_value if 'term_type_duration_col' in locals() else '',
                        'lpr_pending_renewal_months': term_type_duration_value if 'term_type_duration_col' in locals() else '',
                        'lpr_status': "pending",
                        'renew_old_policy': 'true'
                    })

        if error == '':
            messages.success(request, 'File Uploaded Successfully. Data will be processed')
        else:
            err_msg = "No policy has been processed, Fix the below issue and re-upload the excel.<br>"
            messages.error(request, err_msg + str(error))

        return redirect('partners:upload_renewal_policy')

    # Partners Obj
    partners_obj = PartnersDAO.get_partners(condition={'partners_status': 'active'})

    template_name = 'partners/upload_renew_policy.html'
    context = {
        'partners_obj': partners_obj,
        'partners_config': partners_config,
    }

    return render(request, template_name, context)
